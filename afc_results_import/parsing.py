"""
afc_results_import.parsing - read a results workbook into plain data. NO DATABASE.

Deliberately pure: bytes in, structures and problems out. Every layout decision, header spelling and
arithmetic warning is decided here and can be tested without an event, a team or a migration.

TWO SHAPES, because external organizers publish two kinds of data (spec section 6):

  SUMMED (a standings graphic, which is the common case)
      TEAM              MATCHES BOOYAH SCORE ELIMS TOTAL POSITION
      ELITE HUNTERS        6       3     47    82   129      1
    One row per team for a whole group. There is no per-match detail behind it and none is invented.

  PER MATCH (published match by match)
      MATCH MAP      TEAM           PLACE KILLS
      1     Bermuda  ELITE HUNTERS    1     14
    Ordinary results; nothing about aggregate rows applies.

HOW THE SHAPE IS DECIDED
    A MATCH column present means per-match, absent means summed. A STAGE column present means one
    long sheet covering several groups, absent means the sheet name identifies the group.

TOTAL IS TAKEN FROM THE FILE, NEVER RECOMPUTED. The published total is the official result, and an
external organizer's scoring rules are not necessarily AFC's. When SCORE + ELIMS does not equal
TOTAL this warns and keeps the file's number. Silently "correcting" an official standing would be
worse than reporting a discrepancy.

Consumed by afc_results_import.services, which resolves competitors and writes rows.
"""
import io
import re

# Header spellings accepted for each logical column. Matching is case-insensitive and ignores
# surrounding whitespace, punctuation and underscores, so "Team Name", "TEAM_NAME" and "team" all
# land on the same column.
_ALIASES = {
    "team":     {"team", "teamname", "name", "competitor"},
    "matches":  {"matches", "matchesplayed", "played", "m"},
    "booyah":   {"booyah", "booyahs", "wins", "win"},
    "score":    {"score", "placementpoints", "placement", "placepts", "pp"},
    "elims":    {"elims", "eliminations", "kills", "elim"},
    "total":    {"total", "totalpoints", "points", "pts"},
    "position": {"position", "pos", "rank", "place", "standing"},
    "match":    {"match", "matchno", "matchnumber", "game", "gameno"},
    "map":      {"map", "mapname"},
    "placement_in_match": {"place", "placement", "pos", "position"},
    "stage":    {"stage", "phase"},
    "group":    {"group", "grp", "lobby"},
    "advanced": {"advanced", "qualified", "qualifies", "through"},
}


class ParseProblem(Exception):
    """The workbook cannot be read at all. Carries a message written for an admin, not a stack
    trace. Raised only for whole-file failures (unreadable bytes, no usable sheet); a bad ROW is
    reported in `problems` and skipped, never raised."""


def _norm_header(value):
    """Fold a header cell to its comparison key: lowercase, letters and digits only."""
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _header_map(row):
    """Map logical column name -> zero-based index, for the header row given.

    A spreadsheet may carry extra columns AFC does not care about (a flag image, a note). Those are
    ignored rather than rejected: refusing a file for having more information than needed would be
    hostile.
    """
    found = {}
    for idx, cell in enumerate(row):
        key = _norm_header(cell)
        if not key:
            continue
        for logical, spellings in _ALIASES.items():
            if key in spellings and logical not in found:
                found[logical] = idx
    return found


def _int(value, default=None):
    """Best-effort whole number from a spreadsheet cell.

    Cells arrive as int, float ("6.0" when the author typed 6), or str ("6", " 6 ", "1,204"). None
    and blank mean absent, which is different from zero and is preserved as such.
    """
    if value is None:
        return default
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def _truthy(value):
    """Whether an ADVANCED-style cell means yes. Blank means no."""
    return str(value or "").strip().lower() in {"y", "yes", "true", "1", "advanced", "qualified"}


def load_workbook_rows(data):
    """Yield (sheet_name, rows) for every visible sheet. `data` is bytes or a file-like object.

    openpyxl is imported LAZILY, matching afc_tournament_and_scrims.views.export_participants: a
    host missing the wheel must fail on THIS feature with a clear message, not at import time in a
    way that takes unrelated endpoints down with it.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment, not logic
        raise ParseProblem(
            "This server cannot read spreadsheets: the openpyxl package is not installed."
        ) from exc

    try:
        stream = io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data
        wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseProblem(
            "That file could not be opened as a spreadsheet. Save it as .xlsx and try again."
        ) from exc

    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        yield ws.title, [list(r) for r in ws.iter_rows(values_only=True)]


def _find_header(rows):
    """Index of the header row, and its column map.

    Scans the first 10 rows rather than assuming row 1: exported standings often carry a title or a
    blank spacer above the real header. Returns (None, {}) when no row looks like a header.
    """
    for i, row in enumerate(rows[:10]):
        mapping = _header_map(row)
        if "team" in mapping and len(mapping) >= 2:
            return i, mapping
    return None, {}


def parse_sheet(sheet_name, rows):
    """Parse one sheet into {"kind", "group", "stage", "rows", "problems"}.

    kind is "summed" or "per_match". `group` falls back to the sheet name, which is why a
    sheet-per-group workbook needs no STAGE/GROUP columns at all.
    """
    problems = []
    header_idx, cols = _find_header(rows)
    if header_idx is None:
        return {
            "kind": None, "group": sheet_name, "stage": None, "rows": [],
            "problems": [
                f"Sheet {sheet_name!r}: no header row found. Expected a row containing a TEAM "
                f"column, plus either MATCH (per-match results) or MATCHES/TOTAL (summed standings)."
            ],
        }

    kind = "per_match" if "match" in cols else "summed"
    out = []

    for n, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue  # blank spacer row

        def cell(logical):
            idx = cols.get(logical)
            return row[idx] if idx is not None and idx < len(row) else None

        team = str(cell("team") or "").strip()
        if not team:
            problems.append(f"Sheet {sheet_name!r} row {n}: no team name, row skipped.")
            continue

        stage = str(cell("stage") or "").strip() or None
        group = str(cell("group") or "").strip() or sheet_name

        if kind == "per_match":
            match_no = _int(cell("match"))
            if match_no is None:
                problems.append(
                    f"Sheet {sheet_name!r} row {n}: MATCH is not a number, row skipped.")
                continue
            out.append({
                "team": team, "stage": stage, "group": group,
                "match": match_no,
                "map": (str(cell("map") or "").strip().lower() or None),
                "placement": _int(cell("placement_in_match")),
                "kills": _int(cell("elims"), 0),
                "row_number": n,
            })
            continue

        matches = _int(cell("matches"))
        total = _int(cell("total"))
        score = _int(cell("score"), 0)
        elims = _int(cell("elims"), 0)

        if matches is None:
            problems.append(
                f"Sheet {sheet_name!r} row {n} ({team}): MATCHES is missing, row skipped. A summed "
                f"row must say how many matches it covers, or the site cannot report matches played."
            )
            continue
        if total is None:
            # Fall back to the components, and SAY SO. Better than refusing a whole standings sheet
            # over one absent column, and the admin can see what was assumed.
            total = score + elims
            problems.append(
                f"Sheet {sheet_name!r} row {n} ({team}): no TOTAL column, using SCORE + ELIMS "
                f"= {total}.")
        elif score + elims != total:
            # NOT corrected. The published total is the official result; an external organizer's
            # scoring rules are not necessarily AFC's.
            problems.append(
                f"Sheet {sheet_name!r} row {n} ({team}): SCORE {score} + ELIMS {elims} = "
                f"{score + elims}, but TOTAL says {total}. Keeping {total} as published.")

        out.append({
            "team": team, "stage": stage, "group": group,
            "matches": matches,
            "booyah": _int(cell("booyah"), 0),
            "score": score,
            "elims": elims,
            "total": total,
            "position": _int(cell("position")),
            "advanced": _truthy(cell("advanced")),
            "row_number": n,
        })

    return {"kind": kind, "group": sheet_name, "stage": None, "rows": out, "problems": problems}


def parse_workbook(data):
    """Parse every sheet. Returns {"sheets": [...], "problems": [...]}.

    A sheet that yields no usable rows is reported and skipped rather than failing the upload: a
    workbook routinely carries a cover sheet or a notes tab alongside the real standings.
    """
    sheets, problems = [], []
    any_rows = False

    for name, rows in load_workbook_rows(data):
        parsed = parse_sheet(name, rows)
        problems.extend(parsed.pop("problems", []))
        if parsed["rows"]:
            any_rows = True
            sheets.append(parsed)

    if not any_rows:
        raise ParseProblem(
            "No results were found in that workbook. Each sheet needs a header row with a TEAM "
            "column, plus MATCH for per-match results or MATCHES and TOTAL for summed standings. "
            + (" ".join(problems[:3]) if problems else "")
        )

    return {"sheets": sheets, "problems": problems}
