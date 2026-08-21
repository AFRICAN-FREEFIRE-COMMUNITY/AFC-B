"""
afc_results_import.tests_endpoints - the HTTP surface.

Drives the real endpoints with the Django test client and a Bearer token, mirroring the idiom in
afc_tournament_and_scrims.tests_name_matching and afc_rankings.test_ghost_claims.

WHAT MATTERS HERE
    * preview must WRITE NOTHING, over HTTP, not just in the service layer
    * the permission gate must actually refuse a stranger, because these endpoints create teams and
      rewrite results
    * the template must come back as a real workbook carrying THIS event's structure, since that is
      the whole point of the recommended path

Run: python manage.py test afc_results_import.tests_endpoints
"""
import datetime
import io
import secrets

from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile

from afc_auth.models import User, SessionToken
from afc_team.models import Team
from afc_tournament_and_scrims.models import (
    Event, Stages, StageGroups, TournamentTeam, TournamentTeamMatchStats,
)

TODAY = datetime.date.today()

GROUP_A = [
    ["TEAM", "MATCHES", "BOOYAH", "SCORE", "ELIMS", "TOTAL", "POSITION"],
    ["ELITE HUNTERS", 6, 3, 47, 82, 129, 1],
    ["LAXUS E-SPORTS", 6, 1, 54, 55, 109, 2],
]


def _xlsx(sheets):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(data, name="results.xlsx"):
    return SimpleUploadedFile(
        name, data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class ResultsImportEndpointTests(TestCase):

    def setUp(self):
        self.admin = User.objects.create(username="ev_admin", email="ea@example.com", role="admin")
        self.stranger = User.objects.create(
            username="stranger", email="st@example.com", role="player")
        self.admin_token = SessionToken.objects.create(
            user=self.admin, token=secrets.token_hex(32)).token
        self.stranger_token = SessionToken.objects.create(
            user=self.stranger, token=secrets.token_hex(32)).token

        self.event = Event.objects.create(
            slug="ffws-endpoint-test",
            competition_type="tournament", participant_type="squad", event_type="internal",
            max_teams_or_players=16, event_name="FFWS Endpoint Test", event_mode="virtual",
            start_date=TODAY, end_date=TODAY,
            registration_open_date=TODAY, registration_end_date=TODAY,
            prizepool="0", event_rules="r", event_status="ongoing",
            registration_link="https://example.com/r", number_of_stages=1,
        )
        self.stage = Stages.objects.create(
            event=self.event, stage_name="Play-ins", start_date=TODAY, end_date=TODAY,
            number_of_groups=1, stage_format="br - normal", teams_qualifying_from_stage=2)
        self.group = StageGroups.objects.create(
            stage=self.stage, group_name="Group A", playing_date=TODAY,
            playing_time=datetime.time(12, 0), teams_qualifying=2, match_count=6, match_maps=[])

    def _auth(self, token):
        return {"HTTP_AUTHORIZATION": f"Bearer {token}"}

    # ── permission ───────────────────────────────────────────────────────────────────────────────

    def test_anonymous_is_refused(self):
        r = self.client.post("/results-import/preview/",
                             {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))})
        self.assertEqual(r.status_code, 401)

    def test_a_player_is_refused(self):
        """These endpoints create competitors and rewrite results, so the gate is the same one that
        governs reorganising an event's seeding."""
        r = self.client.post(
            "/results-import/preview/",
            {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.stranger_token))
        self.assertEqual(r.status_code, 403)

    # ── preview ──────────────────────────────────────────────────────────────────────────────────

    def test_preview_reports_without_writing(self):
        before = (TournamentTeam.objects.count(), TournamentTeamMatchStats.objects.count())

        r = self.client.post(
            "/results-import/preview/",
            {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(body["preview"]["total_rows"], 2)
        self.assertEqual(body["preview"]["to_create"], 2)
        self.assertEqual(
            (TournamentTeam.objects.count(), TournamentTeamMatchStats.objects.count()), before)

    def test_an_unreadable_file_is_refused_with_a_sentence(self):
        r = self.client.post(
            "/results-import/preview/",
            {"slug": self.event.slug, "file": _upload(b"this is not a spreadsheet", "notes.txt")},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 400)
        self.assertIn("spreadsheet", r.json()["message"].lower())

    def test_a_missing_file_says_so(self):
        r = self.client.post("/results-import/preview/", {"slug": self.event.slug},
                             **self._auth(self.admin_token))
        self.assertEqual(r.status_code, 400)
        self.assertIn("file", r.json()["message"].lower())

    # ── commit ───────────────────────────────────────────────────────────────────────────────────

    def test_commit_writes_the_results(self):
        r = self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 200)
        summary = r.json()["summary"]
        self.assertEqual(summary["stats_rows"], 2)
        self.assertEqual(summary["created_ghosts"], 2)

        row = TournamentTeamMatchStats.objects.get(
            tournament_team__ghost_team__team_name="ELITE HUNTERS")
        self.assertTrue(row.is_aggregate)
        self.assertEqual((row.matches_counted, row.total_points), (6, 129))

    def test_commit_twice_does_not_double_the_results(self):
        payload = {"slug": self.event.slug}
        self.client.post("/results-import/commit/",
                         dict(payload, file=_upload(_xlsx({"Group A": GROUP_A}))),
                         **self._auth(self.admin_token))
        self.client.post("/results-import/commit/",
                         dict(payload, file=_upload(_xlsx({"Group A": GROUP_A}))),
                         **self._auth(self.admin_token))

        self.assertEqual(
            TournamentTeamMatchStats.objects.filter(match__group=self.group).count(), 2)

    # ── pair ─────────────────────────────────────────────────────────────────────────────────────

    def test_pairing_a_name_to_a_real_team(self):
        """The correction tool for a renamed club or a bad match. Not a merge, not a claim."""
        self.client.post("/results-import/commit/",
                         {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
                         **self._auth(self.admin_token))
        real = Team.objects.create(
            team_name="Elite Hunters Official", join_settings="open",
            team_creator=self.admin, team_owner=self.admin)
        tt = TournamentTeam.objects.create(event=self.event, team=real)

        r = self.client.post(
            "/results-import/pair/",
            {"slug": self.event.slug, "source_name": "ELITE HUNTERS",
             "tournament_team_id": tt.pk},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 200)
        self.assertIn("Elite Hunters Official", r.json()["message"])

    def test_pairing_refuses_a_competitor_from_another_event(self):
        other = Event.objects.create(
            slug="other-event",
            competition_type="tournament", participant_type="squad", event_type="internal",
            max_teams_or_players=4, event_name="Other", event_mode="virtual",
            start_date=TODAY, end_date=TODAY,
            registration_open_date=TODAY, registration_end_date=TODAY,
            prizepool="0", event_rules="r", event_status="ongoing",
            registration_link="https://example.com/o", number_of_stages=1)
        team = Team.objects.create(
            team_name="Elsewhere", join_settings="open",
            team_creator=self.admin, team_owner=self.admin)
        foreign_tt = TournamentTeam.objects.create(event=other, team=team)

        r = self.client.post(
            "/results-import/pair/",
            {"slug": self.event.slug, "source_name": "ELITE HUNTERS",
             "tournament_team_id": foreign_tt.pk},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 400)

    # ── template ─────────────────────────────────────────────────────────────────────────────────

    def test_template_carries_this_events_structure(self):
        """The recommended path. The site writes the headers and the team names, so neither can be
        wrong when the file comes back."""
        team = Team.objects.create(
            team_name="Berserk Generation", join_settings="open",
            team_creator=self.admin, team_owner=self.admin)
        tt = TournamentTeam.objects.create(event=self.event, team=team)
        from afc_tournament_and_scrims.models import StageGroupCompetitor
        StageGroupCompetitor.objects.create(stage_group=self.group, tournament_team=tt)

        r = self.client.get(f"/results-import/template/?slug={self.event.slug}",
                            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        sheet = wb["Play-ins - Group A"]
        rows = list(sheet.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "TEAM")
        self.assertEqual(rows[1][0], "Berserk Generation")

    def test_template_round_trips_back_through_the_importer(self):
        """The point of the template: fill it in and it imports cleanly, because the site wrote both
        the headers and the names."""
        team = Team.objects.create(
            team_name="Berserk Generation", join_settings="open",
            team_creator=self.admin, team_owner=self.admin)
        tt = TournamentTeam.objects.create(event=self.event, team=team)
        from afc_tournament_and_scrims.models import StageGroupCompetitor
        StageGroupCompetitor.objects.create(stage_group=self.group, tournament_team=tt)

        tmpl = self.client.get(f"/results-import/template/?slug={self.event.slug}",
                               **self._auth(self.admin_token)).content

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(tmpl))
        ws = wb["Play-ins - Group A"]
        ws.cell(row=2, column=2, value=6)    # MATCHES
        ws.cell(row=2, column=3, value=2)    # BOOYAH
        ws.cell(row=2, column=4, value=30)   # SCORE
        ws.cell(row=2, column=5, value=40)   # ELIMS
        ws.cell(row=2, column=6, value=70)   # TOTAL
        ws.cell(row=2, column=7, value=1)    # POSITION
        buf = io.BytesIO()
        wb.save(buf)

        r = self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "file": _upload(buf.getvalue())},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 200)
        row = TournamentTeamMatchStats.objects.get(tournament_team=tt)
        self.assertEqual((row.matches_counted, row.booyah_count, row.total_points), (6, 2, 70))


class ImportFailSafeTests(ResultsImportEndpointTests):
    """GAP 2 + GAP 3: what a commit decides on the admin's behalf, and what it refuses.

    Inherits the fixture above so the workbook, event and tokens are identical to the commit tests.
    """

    def test_a_first_import_does_NOT_feed_the_rankings_ladder(self):
        """EventCountingControl's rule is "no row => everything counts", so an import used to reach
        the official ladder unless somebody remembered to switch it off. It must fail the safe way:
        an admin turns it ON deliberately, not OFF in a hurry."""
        from afc_rankings.models import EventCountingControl

        self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.admin_token))

        control = EventCountingControl.objects.get(event=self.event)
        self.assertFalse(control.counts_toward_rankings)

    def test_a_first_import_PINS_the_tier(self):
        """Tier is the WEIGHT aggregation applies to results, and the automatic classifier derives
        it from the prize pool, which for an imported event is whatever somebody typed."""
        self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.admin_token))

        self.event.refresh_from_db()
        self.assertTrue(self.event.tier_overridden)

    def test_a_RE_import_does_not_undo_an_admin_who_switched_rankings_on(self):
        """The fail-safe is for the FIRST import only. Re-running a corrected file must not silently
        reverse a decision the admin has since made."""
        from afc_rankings.models import EventCountingControl

        payload = {"slug": self.event.slug}
        self.client.post("/results-import/commit/",
                         dict(payload, file=_upload(_xlsx({"Group A": GROUP_A}))),
                         **self._auth(self.admin_token))
        control = EventCountingControl.objects.get(event=self.event)
        control.counts_toward_rankings = True
        control.save(update_fields=["counts_toward_rankings"])

        self.client.post("/results-import/commit/",
                         dict(payload, file=_upload(_xlsx({"Group A": GROUP_A}))),
                         **self._auth(self.admin_token))

        control.refresh_from_db()
        self.assertTrue(control.counts_toward_rankings)

    def test_per_player_import_is_refused_with_a_reason(self):
        """GAP 3. team_scores_only used to be accepted as false while nothing wrote a per-player
        row, so the API promised an option that did not exist."""
        r = self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "team_scores_only": "false",
             "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 400, r.content[:300])
        self.assertIn("team scores only", r.json()["message"].lower())

    def test_the_default_still_imports_team_scores(self):
        r = self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "file": _upload(_xlsx({"Group A": GROUP_A}))},
            **self._auth(self.admin_token))

        self.assertEqual(r.status_code, 200, r.content[:300])
        self.assertEqual(r.json()["summary"]["stats_rows"], 2)


class TemplateShapeTests(ResultsImportEndpointTests):
    """A template per data shape, because which shape you have decides what the results can be
    used for: only a per-match import can count towards the rankings."""

    def test_the_per_match_template_uses_headers_the_parser_reads_as_per_match(self):
        """The value of a generated template is that the SITE wrote the headers, so a round trip is
        the only test that proves it. PLACEMENT would be read as the summed score column too, which
        is why the sheet says PLACE."""
        import openpyxl
        from afc_results_import.parsing import parse_sheet

        r = self.client.get(
            f"/results-import/template/?slug={self.event.slug}&kind=per_match",
            **self._auth(self.admin_token))
        self.assertEqual(r.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb[wb.sheetnames[0]]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        parsed = parse_sheet(ws.title, rows)

        self.assertEqual(parsed["kind"], "per_match")

    def test_the_default_template_is_still_the_summed_one(self):
        import openpyxl
        from afc_results_import.parsing import parse_sheet

        r = self.client.get(f"/results-import/template/?slug={self.event.slug}",
                            **self._auth(self.admin_token))

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb[wb.sheetnames[0]]
        parsed = parse_sheet(ws.title, [list(row) for row in ws.iter_rows(values_only=True)])
        self.assertEqual(parsed["kind"], "summed")

    def test_an_unknown_kind_is_refused(self):
        r = self.client.get(
            f"/results-import/template/?slug={self.event.slug}&kind=telepathy",
            **self._auth(self.admin_token))
        self.assertEqual(r.status_code, 400)

    def test_a_per_match_template_round_trips_and_CAN_then_count_for_rankings(self):
        """The whole point of offering this shape: fill it in, import it, and the event is
        eligible for the rankings because every finish is present."""
        import openpyxl
        from afc_rankings.models import EventCountingControl

        tmpl = self.client.get(
            f"/results-import/template/?slug={self.event.slug}&kind=per_match",
            **self._auth(self.admin_token)).content
        wb = openpyxl.load_workbook(io.BytesIO(tmpl))
        ws = wb[wb.sheetnames[0]]
        # Type a team name and a real map result, which is what an admin does with a template for
        # an event whose competitors were never entered on AFC (the sheet then carries the header
        # only). The importer creates the competitor as an unclaimed ghost.
        ws.cell(row=2, column=1, value="ELITE HUNTERS")
        ws.cell(row=2, column=2, value=1)    # MATCH
        ws.cell(row=2, column=3, value="bermuda")
        ws.cell(row=2, column=4, value=1)    # PLACE
        ws.cell(row=2, column=5, value=9)    # ELIMS
        buf = io.BytesIO()
        wb.save(buf)

        r = self.client.post(
            "/results-import/commit/",
            {"slug": self.event.slug, "file": _upload(buf.getvalue())},
            **self._auth(self.admin_token))
        self.assertEqual(r.status_code, 200, r.content[:300])

        on = self.client.post(
            "/results-import/settings/",
            {"slug": self.event.slug, "counts_toward_rankings": True},
            content_type="application/json", **self._auth(self.admin_token))

        self.assertEqual(on.status_code, 200, on.content[:300])
        self.assertTrue(
            EventCountingControl.objects.get(event=self.event).counts_toward_rankings)
