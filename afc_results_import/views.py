"""
afc_results_import.views - the HTTP surface for importing an external tournament.

FOUR ENDPOINTS, and the shape of them IS the safety design:

    GET  results-import/template/       download a workbook pre-filled with THIS event's structure
    POST results-import/preview/        upload + parse + resolve, WRITES NOTHING, returns a report
    POST results-import/commit/         write it, in one transaction, after the admin has looked
    POST results-import/pair/           record what a name in the file means (the pairing tool)

Preview and commit are separate on purpose. A results workbook is typed by hand from somebody else's
standings graphic, so it WILL contain surprises: a team AFC has never seen, a name spelled three
ways, a total that does not add up. Preview turns every one of those into a line an admin reads
before anything is written. A bad file produces a report, not a half-imported event.

AUTH mirrors afc_tournament_and_scrims.seeding_management._seeding_gate, so "who may reorganise an
event's seeding" and "who may import its results" are the same people: AFC event admins, the event's
own creator, and organizers holding can_manage_registrations on the owning org. Imported lazily for
the reason that module documents: seeding_management imports afc_auth.views, so a top-level import
here would cycle.

CONSUMED BY: the admin Results Import screen (frontend app/(a)/a/events/[slug]/import).
"""
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from afc_auth.views import validate_token
from afc_tournament_and_scrims.models import Event, StageGroups

from .models import ExternalResultTeamAlias, ResultsImport
from .parsing import ParseProblem
from .services import build_preview, commit_import, norm_team_name

# An upload is a spreadsheet typed by a person, not a data feed. Ten megabytes is far more than any
# realistic standings workbook (the full FFWS structure is a few hundred rows) and small enough that
# a mistaken upload of something else fails fast instead of tying up a worker.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


def _gate(request, event):
    """(user, error_response). One place decides both identity and permission."""
    header = request.headers.get("Authorization") or ""
    if not header.startswith("Bearer "):
        return None, Response({"message": "You need to be signed in to do this."},
                              status=status.HTTP_401_UNAUTHORIZED)
    user = validate_token(header.split(" ", 1)[1])
    if not user:
        return None, Response({"message": "Invalid or expired session token."},
                              status=status.HTTP_401_UNAUTHORIZED)

    # Lazy: seeding_management imports afc_auth.views, so a module-level import cycles.
    from afc_tournament_and_scrims.seeding_management import _seeding_gate
    if not _seeding_gate(user, event):
        return None, Response(
            {"message": "You do not have permission to import results for this event."},
            status=status.HTTP_403_FORBIDDEN)
    return user, None


def _reject_per_player(request):
    """Refuse a request asking for per-player import; None when the request is fine.

    team_scores_only is TRUE and cannot currently be anything else, and the endpoints say so rather
    than accepting a value they will not honour (owner 2026-08-21). The field used to be settable on
    preview while nothing anywhere wrote a single per-player row, so passing false changed NOTHING
    and the API silently promised an option that did not exist.

    Per-player import being unavailable is not an oversight: TournamentPlayerMatchStats.player is a
    foreign key to a real User, and an external tournament has no AFC accounts for its players (FFWS
    Play-ins Phase 1 alone is roughly 720 of them). Inventing those accounts would be far worse than
    having no per-player data, so team-only scoring is the answer rather than a limitation to route
    around. Supporting false means importing player identities first, which is its own feature.

    Checked on BOTH preview and commit: commit accepts a file directly without a prior preview, so
    guarding only the preview would leave the door open on the path that actually writes.
    """
    if str(request.data.get("team_scores_only", "true")).lower() == "false":
        return Response(
            {"message": "Per-player results cannot be imported. A player row needs a real AFC "
                        "account, and an external tournament has none, so an import records team "
                        "scores only. Leave team_scores_only unset or true."},
            status=status.HTTP_400_BAD_REQUEST)
    return None


def _event_or_404(request):
    """Resolve the event from slug or event_id, accepting either in body or query string.

    Both are accepted because the admin screen holds a slug (it routes by slug) while scripts and
    the shell hold an id. Refusing one of them would be arbitrary.
    """
    src = request.data if hasattr(request, "data") and request.data else request.query_params
    slug = (src.get("slug") or "").strip()
    if slug:
        return get_object_or_404(Event, slug=slug)
    return get_object_or_404(Event, pk=src.get("event_id"))


def _read_upload(request):
    """(bytes, error_response). Rejects an absent or oversized file with a sentence, not a stack."""
    f = request.FILES.get("file")
    if f is None:
        return None, Response(
            {"message": "Attach the results workbook as the 'file' field."},
            status=status.HTTP_400_BAD_REQUEST)
    if f.size and f.size > MAX_UPLOAD_BYTES:
        return None, Response(
            {"message": f"That file is {f.size // (1024 * 1024)}MB. The limit is "
                        f"{MAX_UPLOAD_BYTES // (1024 * 1024)}MB."},
            status=status.HTTP_400_BAD_REQUEST)
    return f.read(), None


@api_view(["POST"])
def preview_results_import(request):
    """POST results-import/preview/ - parse and resolve a workbook. WRITES NOTHING.

    Request:  multipart, file=<xlsx>, plus slug or event_id.
    Response: 200 {import_id, preview:{sheets, problems, total_rows, matched, to_create}}
              400 unreadable workbook, with the reason
              403 not permitted for this event

    The `preview` payload is what the admin screen renders: per sheet, whether it is summed or
    per-match, how many rows, and for every competitor whether it matched an existing team, matched
    an existing ghost, or will be created. Near-miss warnings ("this will be created but AFC already
    has something similar") appear in `problems`, because that is the case a person must decide.
    """
    event = _event_or_404(request)
    user, err = _gate(request, event)
    if err:
        return err
    data, err = _read_upload(request)
    if err:
        return err

    err = _reject_per_player(request)
    if err:
        return err

    imp = ResultsImport.objects.create(
        event=event, uploaded_by=user,
        source_filename=(request.FILES["file"].name or "")[:255],
        team_scores_only=True,
    )
    try:
        preview = build_preview(event, data)
    except ParseProblem as exc:
        imp.status = "failed"
        imp.preview = {"problems": [str(exc)]}
        imp.save(update_fields=["status", "preview"])
        return Response({"message": str(exc), "import_id": imp.pk},
                        status=status.HTTP_400_BAD_REQUEST)

    imp.preview = preview
    imp.save(update_fields=["preview"])
    return Response({"import_id": imp.pk, "preview": preview})


@api_view(["POST"])
def commit_results_import(request):
    """POST results-import/commit/ - write a previewed workbook into the event.

    Request:  multipart, file=<xlsx>, slug or event_id, optional import_id.
    Response: 200 {import_id, summary:{groups, created_ghosts, stats_rows, replaced_rows,
                   unmatched_sheets}}
              400 unreadable workbook

    THE FILE IS SENT AGAIN rather than stashed between the two calls. Storing an uploaded blob
    server-side to replay later means a second place for it to go stale or leak, and the admin
    screen already has the bytes in the browser. Re-parsing also means the commit acts on exactly
    what is being confirmed.

    Re-running the same import REPLACES rather than appends, scoped to rows this importer wrote, so
    a hand-entered or log-uploaded result in the same group is never destroyed.
    """
    event = _event_or_404(request)
    user, err = _gate(request, event)
    if err:
        return err
    data, err = _read_upload(request)
    if err:
        return err

    imp = None
    err = _reject_per_player(request)
    if err:
        return err

    if request.data.get("import_id"):
        imp = ResultsImport.objects.filter(pk=request.data["import_id"], event=event).first()
    if imp is None:
        imp = ResultsImport.objects.create(
            event=event, uploaded_by=user,
            source_filename=(request.FILES["file"].name or "")[:255])

    try:
        summary = commit_import(imp, data, actor=user)
    except ParseProblem as exc:
        imp.status = "failed"
        imp.preview = {"problems": [str(exc)]}
        imp.save(update_fields=["status", "preview"])
        return Response({"message": str(exc), "import_id": imp.pk},
                        status=status.HTTP_400_BAD_REQUEST)

    return Response({"import_id": imp.pk, "summary": summary})


@api_view(["POST"])
def pair_result_team(request):
    """POST results-import/pair/ - say what a name in the file MEANS.

    Request:  {slug|event_id, source_name, tournament_team_id}
    Response: 200 {message, alias_id}

    This is the correction tool for a renamed club or a name the matcher read wrongly. It is NOT a
    team merge and NOT a ghost claim: it records AFC's reading of one file, so fixing it is a single
    row update followed by a re-import, with no rosters moved and no history rewritten.

    A CLAIM is the other tool and they are not interchangeable. A claim is a real team asserting
    ownership of a ghost's ranked history, goes through afc_rankings.admin_ghost approval, and moves
    points. Using a claim to fix a typo would be wrong; using an alias to hand a team its history
    would bypass an approval that exists on purpose.
    """
    event = _event_or_404(request)
    user, err = _gate(request, event)
    if err:
        return err

    source_name = (request.data.get("source_name") or "").strip()
    tt_id = request.data.get("tournament_team_id")
    if not source_name or not tt_id:
        return Response(
            {"message": "source_name and tournament_team_id are both required."},
            status=status.HTTP_400_BAD_REQUEST)

    from afc_tournament_and_scrims.models import TournamentTeam
    tt = TournamentTeam.objects.filter(pk=tt_id, event=event).first()
    if tt is None:
        return Response(
            {"message": "That competitor is not registered to this event."},
            status=status.HTTP_400_BAD_REQUEST)

    alias, _ = ExternalResultTeamAlias.objects.update_or_create(
        event=event, normalized_name=norm_team_name(source_name),
        defaults={"source_name": source_name, "tournament_team": tt,
                  "resolution": "manually_paired", "resolved_by": user},
    )
    return Response({
        "message": f"{source_name} will now import as {tt.display_name}. "
                   f"Re-run the import to apply it.",
        "alias_id": alias.pk,
    })


@api_view(["GET"])
def results_import_template(request):
    """GET results-import/template/?slug=... - a workbook pre-filled with this event's structure.

    One sheet per group, already carrying that group's registered competitors in column A and the
    result columns blank. This is the recommended path precisely because it removes two whole
    classes of failure by construction: a header the parser does not recognise, and a team name that
    matches nothing because of a spelling difference. The site wrote both, so both are correct.

    A group with no registered competitors still gets its sheet, with the header only, so an admin
    can type the names in for an event whose teams were never entered on AFC.

    openpyxl is imported LAZILY, matching afc_tournament_and_scrims.views.export_participants: a
    host missing the wheel must fail on THIS endpoint with a clear message rather than at import
    time in a way that takes unrelated endpoints down with it.
    """
    event = _event_or_404(request)
    user, err = _gate(request, event)
    if err:
        return err

    try:
        import openpyxl
    except ImportError:
        return Response(
            {"message": "This server cannot generate spreadsheets: openpyxl is not installed."},
            status=status.HTTP_503_SERVICE_UNAVAILABLE)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    groups = (StageGroups.objects.filter(stage__event=event)
              .select_related("stage").order_by("stage__stage_order", "stage_id", "group_name"))

    for g in groups:
        # Excel sheet titles are capped at 31 characters and cannot contain : \ / ? * [ ]
        title = f"{g.stage.stage_name} - {g.group_name}"
        for ch in ':\\/?*[]':
            title = title.replace(ch, " ")
        ws = wb.create_sheet(title=title[:31])
        ws.append(["TEAM", "MATCHES", "BOOYAH", "SCORE", "ELIMS", "TOTAL", "POSITION"])
        for comp in (g.competitors.select_related("tournament_team__team",
                                                  "tournament_team__ghost_team")
                     .filter(tournament_team__isnull=False)):
            ws.append([comp.tournament_team.display_name, None, None, None, None, None, None])

    if not groups:
        # Better than an empty file: tell the admin what to do about it, in the file itself.
        ws = wb.create_sheet(title="No groups yet")
        ws.append(["This event has no stages or groups yet."])
        ws.append(["Create the structure first, then download this template again."])

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{event.slug or event.pk}-results.xlsx"'
    return resp


# ── GET / POST results-import/settings/ ────────────────────────────────────────────────────────
# The four decisions an admin makes about an imported event, in one place because they are not
# independent and a screen that implies otherwise misleads. Two are about a team's PROFILE, two are
# about the RANKINGS ladder.
#
# WHY THIS EXISTS: the two profile fields shipped enforced but unwritable. Event.
# imported_results_visible_on_profiles and .imported_results_count_in_profile_stats are both read in
# afc_team/views.py, both default False, and nothing anywhere could set them, so an imported event
# was permanently invisible on every team profile and no admin could change it without a shell.
#
# CONSUMED BY: the Results Import tab (frontend app/(a)/a/events/[slug]/edit ResultsImportTab).
_TIERS = {"tier_1", "tier_2", "tier_3"}


def _settings_payload(event):
    """The current state of all four switches, plus what the rankings one resolves to today."""
    from afc_rankings.models import EventCountingControl
    control = EventCountingControl.objects.filter(event=event).first()
    return {
        "slug": event.slug,
        "results_imported_at": event.results_imported_at,
        "visible_on_profiles": event.imported_results_visible_on_profiles,
        "count_in_profile_stats": event.imported_results_count_in_profile_stats,
        # No control row means "everything counts" (EventCountingControl's own rule), so report the
        # EFFECTIVE answer rather than the raw absence of a row.
        "counts_toward_rankings": control.counts_toward_rankings if control else True,
        "tournament_tier": event.tournament_tier,
        "tier_overridden": event.tier_overridden,
    }


@api_view(["GET", "POST"])
def results_import_settings(request):
    """GET returns the four switches; POST updates any subset of them.

    Body (POST): slug (required) plus any of visible_on_profiles, count_in_profile_stats,
    counts_toward_rankings (bool), tournament_tier ("tier_1" | "tier_2" | "tier_3").

    AUTH: the import gate for the two PROFILE switches (whoever may import may decide how the
    import presents). The two RANKINGS switches additionally require an AFC event admin, because
    they change points on a public ladder for teams who have nothing to do with this event, and
    tier is the WEIGHT aggregation applies, not a label.
    """
    slug = request.data.get("slug") or request.query_params.get("slug")
    if not slug:
        return Response({"message": "slug is required."}, status=status.HTTP_400_BAD_REQUEST)
    event = get_object_or_404(Event, slug=slug)
    user, err = _gate(request, event)
    if err:
        return err

    if request.method == "GET":
        return Response(_settings_payload(event))

    from afc_tournament_and_scrims.views import _is_event_admin

    def _flag(name):
        """Tri-state: None when the caller did not mention the field, else a real bool."""
        if name not in request.data:
            return None
        return str(request.data.get(name)).lower() in ("1", "true", "yes", "on")

    visible = _flag("visible_on_profiles")
    counts_stats = _flag("count_in_profile_stats")
    rankings = _flag("counts_toward_rankings")
    tier = request.data.get("tournament_tier")

    if (rankings is not None or tier) and not _is_event_admin(user):
        return Response(
            {"message": "Only an AFC event admin can change what an event contributes to the "
                        "rankings, or its tier."},
            status=status.HTTP_403_FORBIDDEN)
    if tier and tier not in _TIERS:
        return Response({"message": f"tournament_tier must be one of {sorted(_TIERS)}."},
                        status=status.HTTP_400_BAD_REQUEST)

    # ── Refuse to switch rankings ON for a SUMMED import (owner 2026-08-21) ─────────────────
    # The rankings engine derives placement points from a per-match FINISH:
    # aggregation._collect_team does sum(engine.placement_points(s.placement, tables)), and
    # placement_points is a lookup keyed by that finish. A summed import has no per-map finishes to
    # look up: commit_import stores placement=None and keeps only the published placement TOTAL.
    #
    # So switching such an event on does not simply work. It contributes the team's KILLS and ZERO
    # placement points, which is not "this event counts" - it is a silently half-counted event
    # changing a public ladder. Measured on the FFWS import: LAXUS E-SPORTS would have contributed
    # 55 kills and 0 placement points against its published 54.
    #
    # Carrying the stored placement_points instead is NOT the fix and must not be done casually:
    # that column holds the SOURCE tournament's scoring (its own points-per-placement ladder), so
    # importing it injects a foreign scoring system into AFC's rankings, which is the opposite of
    # what the admin scoring config exists to guarantee. engine.placement_points says as much:
    # "callers must NOT trust any legacy placement_points column".
    #
    # Per-match imported sheets are unaffected: those DO store a real placement per map and score
    # exactly like any AFC match.
    if rankings:
        from afc_tournament_and_scrims.models import TournamentTeamMatchStats
        summed = TournamentTeamMatchStats.objects.filter(
            match__group__stage__event=event, is_aggregate=True).exists()
        if summed:
            return Response(
                {"message":
                    "These results are summed standings, not match-by-match, so they cannot count "
                    "towards the rankings. The ranking rules award placement points from a team's "
                    "finish in each map, and a summed import records only the published totals, so "
                    "counting it would credit the kills and none of the placement points. Import "
                    "match-by-match results for this event if it needs to count."},
                status=status.HTTP_400_BAD_REQUEST)

    changed = []
    if visible is not None:
        event.imported_results_visible_on_profiles = visible
        changed.append("imported_results_visible_on_profiles")
    if counts_stats is not None:
        event.imported_results_count_in_profile_stats = counts_stats
        changed.append("imported_results_count_in_profile_stats")
    if tier:
        event.tournament_tier = tier
        # A hand-picked tier is a lock, the same way a head admin's manual tier is: the automatic
        # classifier must not re-derive it from a prize pool nobody imported.
        event.tier_overridden = True
        changed += ["tournament_tier", "tier_overridden"]
    if changed:
        event.save(update_fields=changed)

    if rankings is not None:
        from afc_rankings.models import EventCountingControl
        control, _ = EventCountingControl.objects.get_or_create(
            event=event, defaults={"counts_toward_rankings": rankings, "updated_by": user})
        if control.counts_toward_rankings != rankings:
            control.counts_toward_rankings = rankings
            control.updated_by = user
            control.save(update_fields=["counts_toward_rankings", "updated_by"])

    return Response(_settings_payload(event))
